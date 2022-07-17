import sys
from abc import ABC, abstractmethod
from ..exceptions import ExpectedException
from ..decorators import cli_func
from ..decorators import cli_entry


class NoReportReader(ExpectedException):
    def __init__(self, p):
        super().__init__(f"error: no test reader {p}")


class ReportReader(ABC):
    def __init__(self, rsltFile):
        self.rsltFile = rsltFile

    # Return the file extension of this type of report (e.g. text is .tsv)
    @abstractmethod
    def reportFileExtension(self):
        pass

    # Read the specified level (variant, gene, etc) from a oakvar report
    # Return a list of the report headers and a list of the row values
    # The bDict parameter indicates whether to return Rows as a list or dictionary
    @abstractmethod
    def readReport(self, test_level, bDict):
        pass


# Derived Report Reader class for reating text reports (-t text)
class TextReportReader(ReportReader):
    def reportFileExtension(self):
        return ".tsv"

    # Based on the level selected, return column headers and row values.
    def readReport(self, test_level, bDict):
        level_hdr = "Report level:"
        level = ""
        headers = None
        rows_list = []
        rows_dict = {}
        with open(self.rsltFile, encoding="latin-1") as f:
            line = f.readline().strip("\n")
            while line:
                # skip comment lines but pull out the report level
                if line.strip().startswith("#"):
                    if level_hdr in line:
                        level = line[line.index(level_hdr) + len(level_hdr) + 1 :]
                    line = f.readline().strip("\n")
                    continue

                # only load the level we are testing
                if level != test_level:
                    line = f.readline().strip("\n")
                    continue

                # load headers for the section
                if headers == None:
                    line2 = f.readline().strip("\n")
                    headers = self.readSectionHeader(line, line2)
                    line = f.readline().strip("\n")
                    continue
                columns = line.split("\t")
                line_id = self.getRowID(headers, columns, level)
                if bDict:
                    rows_dict[line_id] = columns
                else:
                    rows_list.append((line_id, columns))
                line = f.readline().strip("\n")
        if bDict:
            return headers, rows_dict
        else:
            return headers, rows_list

    # Read the two report header columns that define the module/column
    # for each data column.  Returned as list of: module|column
    def readSectionHeader(self, line1, line2):
        cols = line1.split("\t")
        headers = []
        current_module = cols[0]
        for col in cols:
            if col != "":
                current_module = col
                headers.append(col + "|")
            else:
                headers.append(current_module + "|")
        cols = line2.split("\t")
        for idx, col in enumerate(cols):
            headers[idx] = headers[idx] + col
        return headers

    # The ID of a result row is used to match key and output.  The ID
    # differs depending on which section of the output is being checked.
    def getRowID(self, headers, columns, level):
        Id = ""
        if level == "variant":
            Id = (
                columns[self.getColPos(headers, "Chrom")]
                + " "
                + columns[self.getColPos(headers, "Position")]
                + " "
                + columns[self.getColPos(headers, "Ref Base")]
                + " "
                + columns[self.getColPos(headers, "Alt Base")]
                + " "
                + columns[self.getColPos(headers, "Tags")]
            )
        if level == "gene":
            pos = self.getColPos(headers, "Hugo")
            if pos == -1:
                pos = self.getColPos(headers, "Gene")
            Id = columns[pos]
        if level == "sample":
            Id = (
                columns[self.getColPos(headers, "UID")]
                + " "
                + columns[self.getColPos(headers, "Sample")]
            )
        if level == "mapping":
            Id = columns[self.getColPos(headers, "Original Line")]
        return Id

    # get the position of a specific output column
    def getColPos(self, headers, col):
        for idx, header in enumerate(headers):
            if col in header:
                return idx
        return -1


# Derived Report Reader class for reating text reports (-t text)
class ExcelReportReader(ReportReader):
    def reportFileExtension(self):
        return ".xlsx"

    # Based on the level selected, return column headers and row values.
    def readReport(self, test_level, bDict):
        from openpyxl import load_workbook

        headers = None
        tabNbr = "Variant"
        if test_level == "gene":
            tabNbr = "Gene"
        elif test_level == "sample":
            tabNbr = "Sample"
        elif test_level == "mapping":
            tabNbr = "Mapping"
        # To open Workbook
        xlxsFile = (
            self.rsltFile if ".xlsx" in self.rsltFile else self.rsltFile + ".xlsx"
        )
        wb = load_workbook(filename=xlxsFile)
        sheet = wb[tabNbr]
        rows_dict = {}
        rows_list = []
        if headers == None:
            headers = self.readSectionHeader(test_level, sheet)
        for i in range(3, sheet.max_row + 1):
            columns = []
            for j in range(1, sheet.max_column + 1):
                columns.append(
                    "" if sheet.cell(i, j).value is None else sheet.cell(i, j).value
                )
            line_id = self.getRowID(headers, columns, test_level)
            if bDict:
                rows_dict[line_id] = columns
            else:
                rows_list.append((line_id, columns))
        if bDict:
            return headers, rows_dict
        else:
            return headers, rows_list

    # Read the two report header columns that define the module/column
    # for each data column.  Returned as list of: module|column
    def readSectionHeader(self, __test_level__, sheet):
        headers = []
        # To open Workbook
        header1 = sheet.cell(1, 1).value
        for i in range(1, sheet.max_column + 1):
            if sheet.cell(1, i).value is not None and sheet.cell(1, i).value != "":
                header1 = sheet.cell(1, i).value
            header2 = sheet.cell(2, i).value
            combinedHeader = header1 + "|" + header2
            headers.append(combinedHeader)
        return headers

    # The ID of a result row is used to match key and output.  The ID
    # differs depending on which section of the output is being checked.
    def getRowID(self, headers, columns, level):
        Id = ""
        if level == "variant":
            Id = (
                columns[self.getColPos(headers, "Chrom")]
                + " "
                + str(int(columns[self.getColPos(headers, "Position")]))
                + " "
                + columns[self.getColPos(headers, "Ref Base")]
                + " "
                + columns[self.getColPos(headers, "Alt Base")]
                + " "
                + columns[self.getColPos(headers, "Tags")]
            )
        if level == "gene":
            pos = self.getColPos(headers, "Hugo")
            if pos == -1:
                pos = self.getColPos(headers, "Gene")
            Id = columns[pos]
        if level == "sample":
            Id = (
                columns[self.getColPos(headers, "UID")]
                + " "
                + columns[self.getColPos(headers, "Sample")]
            )
        if level == "mapping":
            Id = columns[self.getColPos(headers, "Original Line")]
        return Id

    # get the position of a specific output column
    def getColPos(self, headers, col):
        for idx, header in enumerate(headers):
            if col in header:
                return idx
        return -1


# Derived Report Reader class for reating text reports (-t text)
class VcfReportReader(ReportReader):
    def reportFileExtension(self):
        return ".vcf"

    # Based on the level selected, return column headers and row values.
    def readReport(self, __test_level__, bDict):
        headers = None
        rows_dict = {}
        rows_list = []
        import vcf

        reader = vcf.Reader(filename=self.rsltFile)
        if headers == None:
            headers = self.readSectionHeader(reader)
        for record in reader:
            columns = []
            columns.append(record.ID)
            lineitems = record.INFO["CRV"][0].split("|")
            i = 1
            line_id = None
            while i < len(headers):
                columns.append(lineitems[i].strip('"'))
                i += 1
                line_id = self.getRowID(headers, lineitems)
            if bDict:
                rows_dict[line_id] = columns
            else:
                rows_list.append((line_id, columns))
        if bDict:
            return headers, rows_dict
        else:
            return headers, rows_list

    # Read the two report header columns that define the module/column
    # for each data column.  Returned as list of: module|column
    def readSectionHeader(self, reader):
        headers = [
            "Variant Annotation|chrom",
            "Variant Annotation|pos",
            "Variant Annotation|ref",
            "Variant Annotation|alt",
        ]
        colList = reader.infos["CRV"][3]
        cols = colList.split("|")
        for col in cols:
            underscorepos = col.find("__")
            equalspos = col.find("=")
            prevheader = col[underscorepos - 4 : underscorepos]
            postheader = col[underscorepos + 2 : len(col)]
            hugopos = postheader.find("hugo")
            if hugopos >= 0:
                postheader = "hugo"
            elif equalspos > 0:
                break
            if prevheader == "base":
                header = "Variant Annotation|" + postheader
            else:
                module = col[0:underscorepos]
                header = module + "|" + postheader
            headers.append(header)
        return headers

    # The ID of a result row is used to match key and output.  The ID
    # differs depending on which section of the output is being checked.
    def getRowID(self, headers, lineitems):
        Id = (
            lineitems[self.getColPos(headers, "Variant Annotation|chrom")].strip('"')
            + " "
            + lineitems[self.getColPos(headers, "Variant Annotation|pos")].strip('"')
            + " "
            + lineitems[self.getColPos(headers, "Variant Annotation|ref")].strip('"')
            + " "
            + lineitems[self.getColPos(headers, "Variant Annotation|alt")].strip('"')
            + " "
            + lineitems[self.getColPos(headers, "Variant Annotation|tags")].strip('"')
        )
        return Id

    # get the position of a specific output column
    def getColPos(self, headers, col):
        for idx, header in enumerate(headers):
            if col in header:
                return idx
        return -1


# Derived Report Reader class for reating text reports (-t text)
class TsvReportReader(ReportReader):
    def reportFileExtension(self):
        return ".variant.tsv"

    # Based on the level selected, return column headers and row values.
    def readReport(self, test_level, bDict):
        level_hdr = "level="
        level = ""
        headers = None
        rows_dict = {}
        rows_list = []
        with open(self.rsltFile, encoding="latin-1") as f:
            line = f.readline().strip("\n")
            while line:
                # skip comment lines but pull out the report level
                if line.strip().startswith("#"):
                    if level_hdr in line:
                        level = line[line.index(level_hdr) + len(level_hdr) :]
                    line = f.readline().strip("\n")
                    continue

                # only load the level we are testing
                if level != test_level:
                    line = f.readline().strip("\n")
                    continue

                # load headers for the section
                if headers == None:
                    headers = self.readSectionHeader(line)
                    line = f.readline().strip("\n")
                    continue

                columns = line.split("\t")
                line_id = self.getRowID(headers, columns, level)
                if bDict:
                    rows_dict[line_id] = columns
                else:
                    rows_list.append((line_id, columns))
                line = f.readline().strip("\n")
        if bDict:
            return headers, rows_dict
        else:
            return headers, rows_list

    # Read the two report header columns that define the module/column
    # for each data column.  Returned as list of: module|column
    def readSectionHeader(self, line):
        cols = line.split("\t")
        headers = []
        for col in cols:
            header = col.replace(".", "|", 1)
            if not "|" in header:
                header = "Variant Annotation|" + header
            headers.append(header)
        return headers

    # The ID of a result row is used to match key and output.  The ID
    # differs depending on which section of the output is being checked.
    def getRowID(self, headers, columns, level):
        Id = ""
        if level == "variant":
            Id = (
                columns[self.getColPos(headers, "chrom")]
                + " "
                + columns[self.getColPos(headers, "pos")]
                + " "
                + columns[self.getColPos(headers, "ref_base")]
                + " "
                + columns[self.getColPos(headers, "alt_base")]
                + " "
                + columns[self.getColPos(headers, "tags")]
            )
        if level == "gene":
            pos = self.getColPos(headers, "Hugo")
            if pos == -1:
                pos = self.getColPos(headers, "Gene")
            Id = columns[pos]
        if level == "sample":
            Id = (
                columns[self.getColPos(headers, "UID")]
                + " "
                + columns[self.getColPos(headers, "Sample")]
            )
        if level == "mapping":
            Id = columns[self.getColPos(headers, "Original Line")]
        return Id

    # get the position of a specific output column
    def getColPos(self, headers, col):
        for idx, header in enumerate(headers):
            if col in header:
                return idx
        return -1


# Derived Report Reader class for reating text reports (-t text)
class CsvReportReader(ReportReader):
    def reportFileExtension(self):
        return ".variant.csv"

    # Based on the level selected, return column headers and row values.
    def readReport(self, test_level, bDict):
        import csv

        level_hdr = "level="
        level = ""
        headers = None
        rows_dict = {}
        rows_list = []
        with open(self.rsltFile, encoding="latin-1") as f:
            rdr = csv.reader(f)
            for row in rdr:
                # skip comment lines but pull out the report level
                if row[0].startswith("#"):
                    if level_hdr in row[0]:
                        hdr_line = row[0]
                        level = hdr_line[hdr_line.index(level_hdr) + len(level_hdr) :]
                    continue

                # only load the level we are testing
                if level != test_level:
                    continue

                # load headers for the section
                if headers == None:
                    headers = self.readSectionHeader(row)
                    continue

                # load headers for the section
                line_id = self.getRowID(headers, row, level)
                if bDict:
                    rows_dict[line_id] = row
                else:
                    rows_list.append((line_id, row))
        if bDict:
            return headers, rows_dict
        else:
            return headers, rows_list

    # Read the two report header column that define the module/column
    # for each data column.  Returned as list of: module|column
    def readSectionHeader(self, row):
        headers = []
        for col in row:
            # tester use '|' and csv uses ':', just switch the character
            header = col.replace(".", "|", 1)
            # tester is expecting base module to be 'Variant Annotation' so switch it.
            if not "|" in header:
                header = "Variant Annotation|" + header

            headers.append(header)
        return headers

    # The ID of a result row is used to match key and output.  The ID
    # differs depending on which section of the output is being checked.
    def getRowID(self, headers, columns, level):
        Id = ""
        if level == "variant":
            Id = (
                columns[self.getColPos(headers, "chrom")]
                + " "
                + columns[self.getColPos(headers, "pos")]
                + " "
                + columns[self.getColPos(headers, "ref_base")]
                + " "
                + columns[self.getColPos(headers, "alt_base")]
                + " "
                + columns[self.getColPos(headers, "tags")]
            )
        if level == "gene":
            pos = self.getColPos(headers, "Hugo")
            if pos == -1:
                pos = self.getColPos(headers, "Gene")
            Id = columns[pos]
        if level == "sample":
            Id = (
                columns[self.getColPos(headers, "UID")]
                + " "
                + columns[self.getColPos(headers, "Sample")]
            )
        if level == "mapping":
            Id = columns[self.getColPos(headers, "Original Line")]
        return Id

    # get the position of a specific output column
    def getColPos(self, headers, col):
        for idx, header in enumerate(headers):
            if col in header:
                return idx
        return -1


# class that actually runs a test of a specific module and then verifies the results.
class Tester:
    def __init__(self, module, args, input_file):
        from os.path import exists, join
        from os import makedirs
        from ..exceptions import ModuleLoadingError
        from ..module.local import get_local_module_info

        self.parms = None
        self.name = None
        self.args = args
        rundir = args.get("rundir")
        self.module_name = None
        if type(module) == str:
            self.module_name = module
            module = get_local_module_info(self.module_name)
        self.module = module
        if module is None:
            raise ModuleLoadingError(self.module_name)
        if not exists(module.directory) or not module.script_exists:
            raise Exception(
                "No runnable module installed at path %s" % module.directory
            )
        self.out_dir = join(rundir, module.name)
        if not exists(self.out_dir):
            makedirs(self.out_dir)
        self.input_file = input_file
        self.input_path = join(module.test_dir, input_file)
        self.key_path = join(module.test_dir, input_file.replace("input", "key"))
        self.parms_path = join(module.test_dir, input_file.replace("input", "parms"))
        log = "test.log"
        if len(input_file.replace("input", "")) > 0:
            log = input_file + ".test.log"
        self.log_path = join(
            self.out_dir, log
        )  # put the output of this program in test.log
        self.output_file = "oc_output"
        self.out_path = join(self.out_dir, self.output_file)
        self.log = open(self.log_path, "w", encoding="UTF-8")
        self.start_time = None
        self.end_time = None
        self.failures = []
        self.test_passed = False
        self.report_type = "text"

    # optionally a test director for a module can have a 'parms' file.  If it does,
    # the test parameters are tab delimited.  Load each parm/value into the parms
    # dictionary.
    def parse_parms(self):
        from os.path import exists

        self.parms = {}
        if exists(self.parms_path):
            with open(self.parms_path) as f:
                line = f.readline().strip("\n")
                while line:
                    parm = line.split("\t")
                    if len(parm) == 2:
                        # put the parameter and value in the parms dictionary
                        self.parms[parm[0]] = parm[1]
                    line = f.readline().strip("\n")

    # function that tests one module
    def run(self):
        from ..module.local import get_local_module_info
        from time import time
        from subprocess import call, STDOUT
        from ..util.util import quiet_print

        input_msg = (
            "" if self.input_file == "input" else self.input_file
        )  # if there is more than one test for the module, include the test file in the log.
        if self.module is None:
            from ..exceptions import ModuleLoadingError

            raise ModuleLoadingError(self.module_name)
        self._report(f"{self.module.name}: started {input_msg}")
        self.start_time = time()
        self.parse_parms()
        if self.parms is None:
            from ..exceptions import SetupError

            raise SetupError(module_name=self.module_name)
        __python_exc__ = sys.executable
        # default is to run 'text' report but it can be overridden in the optional parms file.
        if "Report_Type" in self.parms:
            self.report_type = self.parms["Report_Type"]
        else:
            self.report_type = "text"
        # Basic oc run command line
        cmd_list = [
            "ov",
            "run",
            self.input_path,
            "-d",
            self.out_dir,
            "-n",
            self.output_file,
            "-t",
            self.report_type,
        ]
        if self.module.type == "annotator":
            cmd_list.extend(["-a", self.module.name])
        elif (
            (self.module.type == "reporter")
            and (get_local_module_info("vest") is not None)
            and (get_local_module_info("cgl") is not None)
        ):
            # when testing reporters, if the vest and cgl modules are installed, include them in the run / report.
            cmd_list.extend(["-a", "vest", "cgl"])
        else:
            cmd_list.extend(["--skip", "annotator"])
        # special case for a few converter modules that need hg19 coordinates
        if self.module.name in [
            "ftdna-converter",
            "ancestrydna-converter",
            "23andme-converter",
        ]:
            cmd_list.extend(["-l", "hg19"])
        else:
            cmd_list.extend(["-l", "hg38"])
        if self.args["to"] == "stdout":
            quiet_print(" ".join(cmd_list), args=self.args)
        exit_code = call(" ".join(cmd_list), shell=True, stdout=self.log, stderr=STDOUT)
        if exit_code != 0:
            self._report(f"{self.module.name}: exit code {exit_code}")
        return exit_code

    def verify(self):
        from ..module.local import get_local_module_info

        if self.module is None:
            from ..exceptions import ModuleLoadingError

            raise ModuleLoadingError(self.module_name)
        self.test_passed = True
        if self.module.type == "annotator":
            self.verify_level(self.module.level, [self.module.title])
        elif self.module.type == "converter":
            self.verify_level("variant", ["Variant Annotation"])
            self.verify_level("sample", ["Variant Annotation"])
            self.verify_level("mapping", ["Variant Annotation"])
        elif self.module.type == "mapper":
            self.verify_level("variant", ["Variant Annotation"])
            self.verify_level("gene", ["Variant Annotation"])
        elif self.module.type == "reporter":
            if self.report_type == "vcf":
                self.verify_level(
                    "variant",
                    [
                        "Variant Annotation",
                        "vest",
                        "cgl",
                        "VEST4",
                        "Cancer Gene Landscape",
                    ],
                )
            else:
                if (get_local_module_info("vest") is not None) and (
                    get_local_module_info("cgl") is not None
                ):
                    self.verify_level(
                        "variant",
                        [
                            "Variant Annotation",
                            "vest",
                            "cgl",
                            "VEST4",
                            "Cancer Gene Landscape",
                        ],
                    )
                    self.verify_level(
                        "gene",
                        [
                            "Variant Annotation",
                            "vest",
                            "cgl",
                            "VEST4",
                            "Cancer Gene Landscape",
                        ],
                    )
                else:
                    self.verify_level("variant", ["Variant Annotation"])
                    self.verify_level("gene", ["Variant Annotation"])

    # See if key and result are floating point numbers.  If so, allow tiny
    # differences.
    def floats_differ(self, str_val1, str_val2):
        try:
            v1 = float(str_val1)
            v2 = float(str_val2)
        except:
            return True
        if abs(v1 - v2) < 0.002:
            return False
        else:
            return True

    # based on the type of report run in this test, create the appropriate type of
    # report reader.
    def create_report_reader(self, type, report_path):
        if type == "text":
            return TextReportReader(report_path)
        elif type == "tsv":
            return TsvReportReader(report_path)
        elif type == "csv":
            return CsvReportReader(report_path)
        elif type == "excel":
            return ExcelReportReader(report_path)
        elif type == "vcf":
            return VcfReportReader(report_path)
        # need to put more parsers here when they are implemented

    # Match the key (expected values) to the text report output.  Generate errors
    # if expected results are not found and fail the test.    Test just the specified
    # level (variant, gene, etc) and specified module's columns
    def verify_level(self, level, module_name):
        # self._report("  Verifying " + level + " level values.")
        if self.module is None:
            from ..exceptions import ModuleLoadingError

            raise ModuleLoadingError(self.module_name)
        key_reader = self.create_report_reader(self.report_type, self.key_path)
        if key_reader is None:
            raise NoReportReader(self.key_path)
        report_extension = key_reader.reportFileExtension()
        report_path = str(self.out_path) + report_extension
        result_reader = self.create_report_reader(self.report_type, report_path)
        if result_reader is None:
            raise NoReportReader(report_path)
        key_header, key_rows = key_reader.readReport(level, False)
        result_header, result_rows = result_reader.readReport(level, True)
        if result_header is None:
            raise NoReportReader(report_path + ":" + level)
        for key in key_rows:
            variant, key_row = key
            if variant not in result_rows:
                self._report(f"{self.module.name}: {variant} did not appear in results")
                self.test_passed = False
                continue
            result = result_rows[variant]
            if key_header is not None:
                for idx, header in enumerate(key_header):
                    # just check the columns from the module we are testing
                    if (
                        (self.getModule(header) not in module_name)
                        or "uid" in header
                        or "UID" in header
                    ):
                        continue
                    if header not in result_header:
                        self._report(
                            f"{self.module.name}: header {header} did not appear in results"
                        )
                        self.test_passed = False
                        continue
                    result_idx = result_header.index(header)
                    if (result[result_idx] != key_row[idx]) and self.floats_differ(
                        result[result_idx], key_row[idx]
                    ):
                        headLabel = header
                        if "|" in header:
                            headLabel = header[header.index("|") + 1 :]
                        self._report(
                            f"{self.module.name}: {variant}/{headLabel}/{key_row[idx]}/{result[result_idx]}"
                        )
                        self.test_passed = False

    # headers are <module name>|<header> - this extracts the module name
    def getModule(self, header):
        return header[: header.index("|")]

    # Write a message to the screen and to the log file.
    def _report(self, s, stdout=False):
        from ..util.util import quiet_print

        self.log.write(s + "\n")
        if stdout:
            quiet_print(s, args=self.args)
        else:
            return s

    # Log success /failure of test.
    def write_results(self, stdout=True):
        if self.module is None:
            from ..exceptions import ModuleLoadingError

            raise ModuleLoadingError(self.module_name)
        if self.start_time is None:
            from ..exceptions import ExpectedException

            raise ExpectedException("start_time does not exist.")
        from time import time

        self.end_time = time()
        elapsed_time = self.end_time - self.start_time
        self._report(f"{self.module.name}: finished in %.2f seconds" % elapsed_time)
        if self.test_passed:
            if stdout:
                self._report(f"{self.module.name}: PASS", stdout=stdout)
            else:
                return "PASS"
        else:
            if stdout:
                self._report(f"{self.module.name}: FAIL", stdout=stdout)
            else:
                return "FAIL"


@cli_entry
def cli_util_test(args):
    args.quiet = False
    args.to = "stdout"
    return test(args)


@cli_func
def test(args, __name__="util test"):
    from os.path import exists
    from os import makedirs
    from ..module.local import get_local_module_types
    from ..module.local import get_local_module_info
    from ..util.util import quiet_print

    rundir = args.get("rundir")
    if rundir is None:
        num = 1
        while True:
            rundir = f"oakvartest_{num}"
            if exists(rundir) == False:
                break
            else:
                num += 1
        args["rundir"] = rundir
    # create run output directory
    if not exists(rundir):
        makedirs(rundir)
    # installed module types
    __module_types__ = get_local_module_types()
    passed = 0
    failed = 0
    modules_failed = []
    module_names = args.get("modules")
    if not module_names:
        from ..exceptions import NoInput

        raise NoInput()
    module_names.sort()
    result = {}
    for module_name in module_names:
        module = get_local_module_info(module_name)
        if module is None:
            continue
        for test_input_file in module.tests:
            tester = Tester(module, args, test_input_file)
            exit_code = tester.run()
            if exit_code == 0:
                tester.verify()
            tester.write_results()
            if tester.test_passed:
                passed += 1
                result[module_name] = {"passed": True, "msg": ""}
            else:
                failed += 1
                fail_msg = module_name + (
                    "" if test_input_file == "input" else " " + test_input_file
                )
                modules_failed.append(fail_msg)
                result[module_name] = {
                    "passed": False,
                    "msg": fail_msg,
                    "log": tester.log_path,
                }
    modules_failed.sort()
    if args.get("to") == "stdout":
        quiet_print(f"passed {passed} failed {failed}", args=args)
    else:
        return {"result": result, "num_passed": passed, "num_failed": failed}


def get_parser_cli_util_test():
    from argparse import ArgumentParser

    parser_cli_util_test = ArgumentParser()
    parser_cli_util_test.add_argument("-d", "--rundir", help="Directory for output")
    parser_cli_util_test.add_argument(
        "-m", "--modules", nargs="+", help="Name of module(s) to test. (e.g. gnomad)"
    )
    parser_cli_util_test.add_argument(
        "-t",
        "--mod_types",
        nargs="+",
        help="Type of module(s) to test (e.g. annotators)",
    )
    parser_cli_util_test.add_argument(
        "--to", default="return", help="stdout to print / return to return"
    )
    parser_cli_util_test.add_argument(
        "--quiet", action="store_true", default=None, help="run quietly"
    )
    parser_cli_util_test.set_defaults(func=cli_util_test)
    return parser_cli_util_test


def main():
    args = get_parser_cli_util_test().parse_args()
    cli_util_test(args)


if __name__ == "__main__":
    main()
